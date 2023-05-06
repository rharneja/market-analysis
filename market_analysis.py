import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os
import calendar
from datetime import datetime, timedelta

def get_data_from_excel_SBI(uploaded_file,month,year):
    wb = load_workbook(uploaded_file)
    ws = wb["Scheme CG"]

    for row in ws.rows:
        for cell in row:
            if cell.value == "Equity Instruments":
                Equity_Instruments_data_start = cell.row + 1
            if cell.value == "Alternate Investments":
                Alternate_Investments_data_start = cell.row + 1

    df = pd.read_excel(uploaded_file,
                                sheet_name = 'Scheme CG',
                                skiprows=Equity_Instruments_data_start - 1,
                                nrows=Alternate_Investments_data_start -Equity_Instruments_data_start -4)
    df = df.drop(['Name of Instruments'], axis=1)
    df = df.rename(columns={'Isin No.': 'ISIN CODE','Mkt_Value':'MARKET VALUE'})
    df.columns = df.columns.str.upper()
    
    # Merge with lookup table to get company names
    lookup_name_with_isin = get_company_name_from_ISIN()
    df = pd.merge(df, lookup_name_with_isin, on='ISIN CODE', how='left')
    
    df = df.applymap(lambda x: x.upper() if type(x) == str else x)
    df['MONTH'] = month
    df['YEAR'] = str(year)
    df['FUND NAME'] = 'SBI'

    return df

def get_data_from_excel_Kotak(uploaded_file,month,year):
    wb = load_workbook(uploaded_file)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # Find the row number of the first cell that contains "Subtotal"
    for row in ws.rows:
        for cell in row:
            if cell.value == "Money Market Instruments":
                rows_till = cell.row
                break # break out of inner loop
        else:
            continue # continue if the inner loop did not break
        break # break out of outer loop

    # Load the data from the Excel file into a DataFrame
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, skiprows=11, nrows=rows_till - 13)
    # Clean up the DataFrame
    df = df.drop(df[(df['Name of the Instrument'] == 'Equity Instruments') | (df['Name of the Instrument'] == '       Shares')].index)
    df.dropna(subset=['Name of the Instrument'],inplace = True)
    df = df[df.filter(regex='^(?!Unnamed)').columns]
    df = df.drop(['Name of the Instrument','Ratings'], axis=1)
    df = df.rename(columns={'ISIN No.': 'ISIN CODE','Industry':'INDUSTRY','Market Value Rs.': 'MARKET VALUE'})
    df.columns = df.columns.str.upper()
    df = df.applymap(lambda x: x.upper() if type(x) == str else x)
    

    # Merge with lookup table to get company names
    lookup_name_with_isin = get_company_name_from_ISIN()
    df = pd.merge(df, lookup_name_with_isin, on='ISIN CODE', how='left')

    # Add additional columns to the DataFrame
    df['MONTH'] = month
    df['YEAR'] = str(year)
    df['FUND NAME'] = 'KOTAK'

    return df

def get_data_from_excel_ICICI(uploaded_file,month,year):
    wb = load_workbook(uploaded_file)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # Find the row number of the first cell that contains "Subtotal"
    for row in ws.rows:
        for cell in row:
            if cell.value == "Subtotal":
                rows_till = cell.row
                break # break out of inner loop
        else:
            continue # continue if the inner loop did not break
        break # break out of outer loop

    # Load the data from the Excel file into a DataFrame
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, skiprows=5, nrows=rows_till - 7)

    # Clean up the DataFrame
    df = df.drop(df[(df['Particulars'] == 'Equity Instruments') | (df['Particulars'] == 'Shares')].index)
    df = df.drop(['Particulars'], axis=1)
    df = df.rename(columns={'ISIN No.': 'ISIN CODE','Industry':'INDUSTRY'})
    df.columns = df.columns.str.upper()
    df = df.applymap(lambda x: x.upper() if type(x) == str else x)
    

    # Merge with lookup table to get company names
    lookup_name_with_isin = get_company_name_from_ISIN()
    df = pd.merge(df, lookup_name_with_isin, on='ISIN CODE', how='left')

    # Add additional columns to the DataFrame
    df['MONTH'] = month
    df['YEAR'] = str(year)
    df['FUND NAME'] = 'ICICI'

    return df

def get_data_from_excel_HDFC(uploaded_file):
    df = pd.DataFrame({})
    return df

def get_data_from_excel_AB(uploaded_file,month,year):
    wb = load_workbook(uploaded_file)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # Find the row number of the first cell that contains "Money Market Instruments:-"
    for row in ws.rows:
        for cell in row:
            if cell.value == "Money Market Instruments:-":
                rows_till = cell.row
                break

    # Load the data from the Excel file into a DataFrame
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, skiprows=5, nrows=rows_till - 10)

    # Clean up the DataFrame
    df = df.drop(['Unnamed: 0', 'Ratings', 'Name of the Instrument'], axis=1)
    df = df.rename(columns={'ISIN No.': 'ISIN CODE','Mkt_Value':'MARKET VALUE','Industry ':'INDUSTRY'})
    df.columns = df.columns.str.upper()
    df = df.applymap(lambda x: x.upper() if type(x) == str else x)

    # Merge with lookup table to get company names
    lookup_name_with_isin = get_company_name_from_ISIN()
    df = pd.merge(df, lookup_name_with_isin, on='ISIN CODE', how='left')

    # Add additional columns to the DataFrame
    df['MONTH'] = month
    df['YEAR'] = str(year)
    df['FUND NAME'] = 'ADITYA BIRLA'

    return df

def get_data_from_excel_UTI(uploaded_file,month,year):
    wb = load_workbook(uploaded_file)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # Find the row number of the first cell that contains "Subtotal"
    for row in ws.rows:
        for cell in row:
            if cell.value == "Debt Instruments -":
                rows_till = cell.row
                break # break out of inner loop
        else:
            continue # continue if the inner loop did not break
        break # break out of outer loop

    # Load the data from the Excel file into a DataFrame
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, skiprows=4, nrows=rows_till - 6)
    # Clean up the DataFrame
    df = df.drop(df[(df['Name of the Instrument'] == 'Equity Instruments -') | (df['Name of the Instrument'] == 'Shares')].index)
    df.dropna(subset=['Name of the Instrument'],inplace = True)
    df = df[df.filter(regex='^(?!Unnamed)').columns]
    df = df.drop(['Name of the Instrument','Rating','Industry Code'], axis=1)
    df = df.rename(columns={'ISIN No.': 'ISIN CODE','Industry Name':'INDUSTRY','Mkt Value': 'MARKET VALUE'})
    df.columns = df.columns.str.upper()
    df = df.applymap(lambda x: x.upper() if type(x) == str else x)
    

    # Merge with lookup table to get company names
    lookup_name_with_isin = get_company_name_from_ISIN()
    df = pd.merge(df, lookup_name_with_isin, on='ISIN CODE', how='left')

    # Add additional columns to the DataFrame
    df['MONTH'] = month
    df['YEAR'] = str(year)
    df['FUND NAME'] = 'UTI'

    return df

def get_data_from_excel_LIC(uploaded_file,month,year):
    wb = load_workbook(uploaded_file)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # Find the row number of the first cell that contains "Subtotal"
    for row in ws.rows:
        for cell in row:
            if cell.value == "Total (A)":
                rows_till = cell.row
                break # break out of inner loop
        else:
            continue # continue if the inner loop did not break
        break # break out of outer loop

    # Load the data from the Excel file into a DataFrame
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, skiprows=7, nrows=rows_till - 9)
    # Clean up the DataFrame
    df = df[df.filter(regex='^(?!Unnamed)').columns]
    df = df.drop(['Security Name','Rating'], axis=1)
    df = df.rename(columns={'ISIN Code': 'ISIN CODE','NAV%':'% OF PORTFOLIO','Units': 'QUANTITY'})
    df.columns = df.columns.str.upper()
    df = df.applymap(lambda x: x.upper() if type(x) == str else x)
    

    # Merge with lookup table to get company names
    lookup_name_with_isin = get_company_name_from_ISIN()
    df = pd.merge(df, lookup_name_with_isin, on='ISIN CODE', how='left')

    # Add additional columns to the DataFrame
    df['MONTH'] = month
    df['YEAR'] = str(year)
    df['FUND NAME'] = 'LIC'

    return df

def get_data_from_excel_MAX(uploaded_file,month,year):
    wb = load_workbook(uploaded_file)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # Find the row number of the first cell that contains "Subtotal"
    for row in ws.rows:
        for cell in row:
            if cell.value == "Subtotal (A)":
                rows_till = cell.row
                break # break out of inner loop
        else:
            continue # continue if the inner loop did not break
        break # break out of outer loop

    # Load the data from the Excel file into a DataFrame
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, skiprows=4, nrows=rows_till - 5)
    # Clean up the DataFrame
    df.dropna(subset=['ISIN No.'],inplace = True)
    df = df[df.filter(regex='^(?!Unnamed)').columns]
    df = df.drop(['Name of the Instrument'], axis=1)
    df = df.rename(columns={'ISIN No.': 'ISIN CODE','Industry ':'Industry'})
    df.columns = df.columns.str.upper()
    df = df.applymap(lambda x: x.upper() if type(x) == str else x)
    

    # Merge with lookup table to get company names
    lookup_name_with_isin = get_company_name_from_ISIN()
    df = pd.merge(df, lookup_name_with_isin, on='ISIN CODE', how='left')

    # Add additional columns to the DataFrame
    df['MONTH'] = month
    df['YEAR'] = str(year)
    df['FUND NAME'] = 'MAX'

    return df

def get_data_from_excel_TATA(uploaded_file,month,year):
    wb = load_workbook(uploaded_file)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # Find the row number of the first cell that contains "Subtotal"
    for row in ws.rows:
        for cell in row:
            if cell.value == "Money Market Instruments":
                rows_till = cell.row
                break # break out of inner loop
        else:
            continue # continue if the inner loop did not break
        break # break out of outer loop

    # Load the data from the Excel file into a DataFrame
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, skiprows=6, nrows=rows_till - 13)
    # Clean up the DataFrame
    df.dropna(subset=['Isin No.'],inplace = True)
    df = df[df.filter(regex='^(?!Unnamed)').columns]
    df = df.drop(['Name of Instruments'], axis=1)
    df = df.rename(columns={'Isin No.': 'ISIN CODE','Mkt_Value': 'MARKET VALUE'})
    df.columns = df.columns.str.upper()
    df = df.applymap(lambda x: x.upper() if type(x) == str else x)
    

    # Merge with lookup table to get company names
    lookup_name_with_isin = get_company_name_from_ISIN()
    df = pd.merge(df, lookup_name_with_isin, on='ISIN CODE', how='left')

    # Add additional columns to the DataFrame
    df['MONTH'] = month
    df['YEAR'] = str(year)
    df['FUND NAME'] = 'TATA'

    return df

@st.cache_data()
def get_company_name_from_ISIN():
    df = pd.read_csv("ind_nifty500list.csv",
                    usecols = ['ISIN Code','Company Name'])
    df.columns = df.columns.str.upper()
    df = df.applymap(lambda x: x.upper() if isinstance(x, str) else x)
    return df

@st.cache_data()
def transform_nps_trust_df(nps_trust_scheme_cg_df,selected_value):
    nps_trust_scheme_cg_df['YEAR-MONTH'] = nps_trust_scheme_cg_df.apply(lambda x: f"{x['YEAR']}-{x['MONTH']:02d}", axis=1)
    nps_pivot = nps_trust_scheme_cg_df.pivot_table(values=selected_value, index='COMPANY NAME', columns=['YEAR','MONTH'],
                                                    aggfunc='sum',fill_value=0)
    nps_pivot.columns = [f"{calendar.month_abbr[m].upper()}-{y}" for y, m in nps_pivot.columns]

    return nps_pivot

@st.cache_data()
def apply_months_filter(nps_trust_scheme_cg_df,slider_value):
    if slider_value == 0:
        return nps_trust_scheme_cg_df
    else:
        filter_date = datetime.now() - timedelta(days=slider_value*30)
        nps_trust_scheme_cg_df['date'] = pd.to_datetime(nps_trust_scheme_cg_df['MONTH'].astype(str) + nps_trust_scheme_cg_df['YEAR'].astype(str), format='%m%Y')
        filtered_data = nps_trust_scheme_cg_df[nps_trust_scheme_cg_df['date'] >= filter_date]
        return filtered_data

def nps_color(df):
    conds = [
        df.gt(df.shift(+1, axis=1)), #is the current column greater than the previous ?
        df.lt(df.shift(+1, axis=1)), #is the current column lower than the previous ?
    ]
    vals = [
        "background-color: lightgreen",
        "background-color: lightcoral",
    ]
    return pd.DataFrame(np.select(conds, vals, default=""),
                        index=df.index, columns=df.columns)


    # create an empty style object
    style = pd.Series(np.nan, index=s.index.astype(float))
    # loop over columns in the dataframe
    for i in range(1, len(s)):
        if isinstance(s, pd.Series) and s[i] < s[i-1]:
            # set the style for the cell to red if the value is less than the previous column's value
            style[i] = 'color: red'
        elif isinstance(s, pd.Series) and s[i] > s[i-1]:
            # set the style for the cell to green if the value is greater than the previous column's value
            style[i] = 'color: green'
        elif isinstance(s, pd.DataFrame) and s.iloc[:, i].lt(s.iloc[:, i-1]).any():
            # set the style for the entire column to red if any value in the column is less than the previous column's value
            style.iloc[:, i] = 'color: red'
        elif isinstance(s, pd.DataFrame) and s.iloc[:, i].gt(s.iloc[:, i-1]).any():
            # set the style for the entire column to green if any value in the column is greater than the previous column's value
            style.iloc[:, i] = 'color: green'
    return style

if 'filter_shares' not in st.session_state:
    st.session_state['filter_shares'] = 'value'


st.set_page_config(page_title = 'Market Analysis',  
                    page_icon = ":bar_chart:",
                    layout = 'wide')
st.header('Market Analysis')
st.subheader('NPS Trust Portfolio')

if os.path.exists('scheme_cg.csv'):
    nps_trust_scheme_cg_df = pd.read_csv('scheme_cg.csv')
    nps_trust_scheme_cg_df.sort_values(['COMPANY NAME', 'YEAR','MONTH'], inplace=True)
else:
    nps_trust_scheme_cg_df = None
    st.info('No NPS Scheme CG Data Available')

nps_upload_form = st.form('Upload NPS Trust Files', clear_on_submit=True)
uploaded_files = nps_upload_form.file_uploader('Upload your files here',accept_multiple_files=True)
submit = nps_upload_form.form_submit_button('Submit', use_container_width = True)
nps_fund_names = {'SBI':get_data_from_excel_SBI,
                    'KOTAK':get_data_from_excel_Kotak,
                    'ICICI':get_data_from_excel_ICICI,
                    'HDFC':get_data_from_excel_HDFC,
                    'AB':get_data_from_excel_AB,
                    'UTI':get_data_from_excel_UTI,
                    'LIC':get_data_from_excel_LIC,
                    'MAX':get_data_from_excel_MAX,
                    'TATA':get_data_from_excel_TATA}
months = {'JAN' : 1, 'FEB' : 2, 'MAR' : 3, 'APR' : 4, 'MAY' : 5, 'JUN' : 6,
         'JUL' : 7, 'AUG' : 8, 'SEP' : 9, 'OCT': 10, 'NOV' : 11, 'DEC' : 12}

if submit:
    for uploaded_file in uploaded_files:
        file_name = uploaded_file.name
        with st.spinner(f'Processing {file_name}'):
            fund_manager, month, year = file_name.split('_')
            year = year.split('.')[0]
            month = months[month.upper()]

            uploaded_data = nps_fund_names[fund_manager.upper()](uploaded_file,month,year)
            if isinstance(nps_trust_scheme_cg_df, pd.DataFrame):
                nps_trust_scheme_cg_df = pd.concat([nps_trust_scheme_cg_df, uploaded_data])
            else:
                nps_trust_scheme_cg_df = uploaded_data
            nps_trust_scheme_cg_df.to_csv('scheme_cg.csv',index = False)
    st.commands.execution_control.rerun()

if st.sidebar.button('Reset'):
    if os.path.exists("scheme_cg.csv"):
        os.remove("scheme_cg.csv")
    st.commands.execution_control.rerun()
#group_by_cols = st.sidebar.multiselect('Select Group by Columns', options = ['YEAR', 'MONTH'], default=['YEAR','MONTH']  )

if isinstance(nps_trust_scheme_cg_df, pd.DataFrame):
    selected_value = st.sidebar.selectbox('Select a Value',options = ['QUANTITY','MARKET VALUE','% OF PORTFOLIO'])
    company_name = st.sidebar.multiselect('Filter Shares',options = nps_trust_scheme_cg_df['COMPANY NAME'].unique(),default=None)
    slider_value = st.sidebar.slider("Select time period", 0, 36, 0, 3, format="%d months", key="slider")
    
    nps_trust_scheme_cg_df = apply_months_filter(nps_trust_scheme_cg_df,slider_value)
    if company_name:
        st.session_state['filter_shares'] = company_name
        nps_trust_scheme_cg_df = nps_trust_scheme_cg_df.query('`COMPANY NAME` == @company_name')
        st.dataframe(nps_trust_scheme_cg_df,use_container_width = True)

    nps_pivot = transform_nps_trust_df(nps_trust_scheme_cg_df,selected_value)
    st.dataframe(nps_pivot.style.apply(nps_color, axis=None),use_container_width = True)

